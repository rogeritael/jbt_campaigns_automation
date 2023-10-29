from openpyxl import Workbook

workbook = Workbook()
new_sheet = workbook.create_sheet('titulo')
workbook.active = new_sheet

current_row = workbook.active.max_row


values = ['23.54', '300', '0.15%']

for index, param in enumerate(values):
    char = chr(index + 65)
    cell = workbook.active[f'{char}{current_row}']
    cell.value = param

workbook.save('./arquivo.xlsx')