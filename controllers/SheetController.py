from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

class SheetController:
    def __init__(self):
        self.workbook = Workbook()
        self.active_sheet = None
        self.current_row = 1
        default_sheet = self.workbook.get_sheet_by_name('Sheet')
        if default_sheet:
            self.workbook.remove(default_sheet)

    def createSheet(self, titulo: str = ''):
        new_sheet = self.workbook.create_sheet(titulo)
        self.workbook.active = new_sheet
        self.active_sheet = new_sheet

        return new_sheet

    def createRow(self, columns: [], highlight: bool = False):
        for index, column in enumerate(columns):

            cell = self.active_sheet[f'{chr(index + 65)}{self.current_row}']
            cell.alignment =  Alignment(vertical='center', horizontal='center')
            cell.font = Font(name='Space Grotesk')
            self.active_sheet.column_dimensions[f'{chr(index + 65)}'].width = 15

            if highlight:
                cell.font = Font(b=True)
            
            cell.value = column
        
        self.current_row += 1
    
    def createHeader(self, value: str, range: int, height: int = 25):
        row = self.current_row
        
        self.active_sheet.merge_cells(f'A{row}:{chr(range + 65)}{row}')
        cell = self.active_sheet[f'A{row}']
        cell.value = value
        cell.alignment =  Alignment(vertical='center', horizontal='center')
        cell.fill = PatternFill('solid', fgColor='4f24ee')
        cell.font = Font(name='Space Grotesk', b=True, color='FFFFFF')
        self.active_sheet.row_dimensions[self.current_row].height = height
        
        self.current_row += 1

    

    def save(self):
        self.workbook.save('./sheets/dados.xlsx')